import sqlite3
import os
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

print("=" * 70)
print("DAY 20 - PEER COMPARISON EXCEL")
print("=" * 70)

os.makedirs("output", exist_ok=True)

conn = sqlite3.connect("db/nifty100.db")

query = """
SELECT
    fr.company_id,
    fr.year,
    fr.sales,
    fr.return_on_equity_pct,
    fr.debt_to_equity,
    fr.free_cash_flow_cr,
    fr.revenue_cagr_5yr,
    fr.pat_cagr_5yr,
    fr.operating_profit_margin_pct,
    fr.interest_coverage,
    fr.asset_turnover,
    fr.dividend_payout_ratio_pct,
    fr.composite_quality_score,
    s.broad_sector,
    s.sub_sector

FROM financial_ratios fr

LEFT JOIN sectors s
ON fr.company_id = s.company_id
"""

df = pd.read_sql(query, conn)
conn.close()

df = df[df["year"] != "TTM"].copy()

df["year_num"] = (
    df["year"]
    .astype(str)
    .str.extract(r"(\d{4})")
    .astype(int)
)

df = df.sort_values(
    ["company_id", "year_num"],
    ascending=[True, False]
)

df = df.drop_duplicates(
    subset="company_id",
    keep="first"
)

df.drop(columns="year_num", inplace=True)

print(f"\nCompanies Loaded : {len(df)}")

print("\nAvailable Peer Groups:")
print(sorted(df["broad_sector"].dropna().unique()))
print("Total:", df["broad_sector"].nunique())

wb = Workbook()

# remove default sheet
wb.remove(wb.active)

green_fill = PatternFill(
    fill_type="solid",
    fgColor="C6EFCE"
)

yellow_fill = PatternFill(
    fill_type="solid",
    fgColor="FFF2CC"
)

red_fill = PatternFill(
    fill_type="solid",
    fgColor="F4CCCC"
)

gold_fill = PatternFill(
    fill_type="solid",
    fgColor="FFD966"
)

header_fill = PatternFill(
    fill_type="solid",
    fgColor="4472C4"
)

header_font = Font(
    bold=True,
    color="FFFFFF"
)

sectors = sorted(df["broad_sector"].dropna().unique())

print("\nPeer Groups Found:")

for s in sectors:
    print(" -", s)
    
for sector in sectors:

    sector_df = (
        df[df["broad_sector"] == sector]
        .copy()
        .sort_values(
            "composite_quality_score",
            ascending=False
        )
    )

    ws = wb.create_sheet(title=sector[:31])

    headers = [
        "Company ID",
        "Sector",
        "Sub Sector",
        "Sales",
        "ROE",
        "Debt/Equity",
        "Free Cash Flow",
        "Revenue CAGR 5Y",
        "PAT CAGR 5Y",
        "OPM",
        "Interest Coverage",
        "Asset Turnover",
        "Dividend Payout",
        "Composite Score"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, (_, row) in enumerate(sector_df.iterrows(), start=2):

        values = [
            row["company_id"],
            row["broad_sector"],
            row["sub_sector"],
            row["sales"],
            row["return_on_equity_pct"],
            row["debt_to_equity"],
            row["free_cash_flow_cr"],
            row["revenue_cagr_5yr"],
            row["pat_cagr_5yr"],
            row["operating_profit_margin_pct"],
            row["interest_coverage"],
            row["asset_turnover"],
            row["dividend_payout_ratio_pct"],
            row["composite_quality_score"]
        ]

        for col_idx, value in enumerate(values, start=1):

            cell = ws.cell(
                row=row_idx,
                column=col_idx
            )

            cell.value = value

            if isinstance(value, (int, float)):

                if value >= 75:
                    cell.fill = green_fill

                elif value >= 25:
                    cell.fill = yellow_fill

                else:
                    cell.fill = red_fill

    median_row = ws.max_row + 2

    ws.cell(
        row=median_row,
        column=1
    ).value = "Median"

    ws.cell(
        row=median_row,
        column=1
    ).fill = gold_fill

    numeric_cols = {
        4: "sales",
        5: "return_on_equity_pct",
        6: "debt_to_equity",
        7: "free_cash_flow_cr",
        8: "revenue_cagr_5yr",
        9: "pat_cagr_5yr",
        10: "operating_profit_margin_pct",
        11: "interest_coverage",
        12: "asset_turnover",
        13: "dividend_payout_ratio_pct",
        14: "composite_quality_score"
    }

    for excel_col, df_col in numeric_cols.items():

        ws.cell(
            row=median_row,
            column=excel_col
        ).value = sector_df[df_col].median()

    for column_cells in ws.columns:

        length = max(
            len(str(cell.value))
            if cell.value is not None else 0
            for cell in column_cells
        )

        ws.column_dimensions[
            get_column_letter(column_cells[0].column)
        ].width = min(length + 3, 25)

output_file = "output/peer_comparison.xlsx"

wb.save(output_file)

print("\n" + "=" * 70)
print("PEER COMPARISON EXCEL GENERATED SUCCESSFULLY")
print("=" * 70)

print(f"\nTotal Peer Groups : {len(sectors)}")

print(f"Output File : {output_file}")

print("\nEach worksheet contains:")
print("✓ Company financial metrics")
print("✓ Sector-wise grouping")
print("✓ Composite quality score")
print("✓ Median summary row")
print("✓ Conditional colour formatting")

print("\nSprint 3 Peer Comparison Report Completed!") 